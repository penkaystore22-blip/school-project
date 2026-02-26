"""
Generate MAGANA CBC School MIS Professional Workbook
Creates all required sheets with headers, pre-populated data — NO sheet protection.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

OUTPUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "MAGANA_CBC_SchoolMIS_Professional.xlsx")

wb = openpyxl.Workbook()

# ── Style definitions ──
hdr_font = Font(name="Segoe UI", bold=True, size=11, color="FFFFFF")
hdr_fill = PatternFill("solid", fgColor="1F3864")
hdr_align = Alignment(horizontal="center", vertical="center")
thin = Side(style="thin", color="B0B0B0")
hdr_border = Border(bottom=thin)

def make_sheet(name, headers, col_widths=None):
    """Create a sheet with styled header row."""
    ws = wb.create_sheet(title=name)
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = hdr_border
    if col_widths:
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    else:
        for i in range(1, len(headers) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 18
    ws.sheet_properties.tabColor = "1F3864"
    ws.freeze_panes = "A2"
    return ws

def seed_data(ws, rows, start_row=2):
    """Write multiple rows of data starting from start_row."""
    for r_idx, row_data in enumerate(rows, start_row):
        for c_idx, val in enumerate(row_data, 1):
            ws.cell(row=r_idx, column=c_idx, value=val)

# ===================== DATA SHEETS =====================

# 1. DB_Students (17 columns matching cSID..cSTRM)
make_sheet("DB_Students", [
    "StudentID", "AdmNo", "FirstName", "LastName", "Gender",
    "DOB", "Class", "ParentName", "Phone", "UPI",
    "BirthCert", "Status", "Allergies", "SpecialNeeds", "BloodGroup",
    "PhotoPath", "Stream"
])

# 2. DB_Classes — All CBC classes with streams
ws_cls = make_sheet("DB_Classes", ["ClassName", "StreamName"])
classes_data = []
cbc_classes = ["PP1", "PP2", "G1", "G2", "G3", "G4", "G5", "G6", "G7", "G8", "G9"]
default_streams = ["A", "B"]
for cls in cbc_classes:
    for strm in default_streams:
        classes_data.append((cls, strm))
seed_data(ws_cls, classes_data)

# 3. DB_Teachers
make_sheet("DB_Teachers", [
    "TeacherID", "TSCNo", "FirstName", "LastName", "Gender",
    "Phone", "Email", "Subject", "Class", "DateJoined", "Status"
])

# 4. DB_Guardians
make_sheet("DB_Guardians", [
    "GuardianID", "StudentID", "Name", "Relationship", "Phone", "Email", "Address"
])

# 5. DB_Attendance
make_sheet("DB_Attendance", [
    "Date", "StudentID", "Class", "Status", "Remarks"
])

# 6. DB_FeeStructure — Pre-populated with typical Kenyan school fees
ws_fee = make_sheet("DB_FeeStructure", ["ID", "FeeType", "Amount"])
fee_data = [
    (1, "Tuition Fee",       12000),
    (2, "Activity Fee",       2500),
    (3, "Exam Fee",           1500),
    (4, "Lunch Program",      4500),
    (5, "Transport (Optional)", 3000),
    (6, "Uniform Levy",       2000),
    (7, "ICT Levy",           1000),
    (8, "Development Fund",   1500),
    (9, "Insurance",           800),
    (10, "Stationery",        1200),
]
seed_data(ws_fee, fee_data)

# 7. DB_FeePayments
make_sheet("DB_FeePayments", [
    "PaymentID", "StudentID", "FeeType", "Amount", "PayMode",
    "ReceiptNo", "Date", "RecordedBy"
])

# 8. DB_Subjects — Pre-populated with full CBC curriculum subjects
ws_subj = make_sheet("DB_Subjects", ["ID", "Level", "SubjectName", "MaxMarks"])
subj_data = [
    # Pre-Primary (PP)
    (1,  "PP", "Language Activities",           100),
    (2,  "PP", "Mathematical Activities",       100),
    (3,  "PP", "Environmental Activities",      100),
    (4,  "PP", "Psychomotor & Creative Arts",   100),
    (5,  "PP", "Religious Education",           100),
    # Lower Primary (Grade 1-3)
    (6,  "Lower Primary", "Kiswahili",                    100),
    (7,  "Lower Primary", "English",                      100),
    (8,  "Lower Primary", "Mathematics",                  100),
    (9,  "Lower Primary", "Environmental Activities",     100),
    (10, "Lower Primary", "Hygiene & Nutrition",          100),
    (11, "Lower Primary", "Religious Education (CRE)",    100),
    (12, "Lower Primary", "Movement & Creative Arts",     100),
    (13, "Lower Primary", "Indigenous Languages",         100),
    # Upper Primary (Grade 4-6)
    (14, "Upper Primary", "English",                      100),
    (15, "Upper Primary", "Kiswahili",                    100),
    (16, "Upper Primary", "Mathematics",                  100),
    (17, "Upper Primary", "Science & Technology",         100),
    (18, "Upper Primary", "Social Studies",               100),
    (19, "Upper Primary", "Religious Education (CRE)",    100),
    (20, "Upper Primary", "Agriculture",                  100),
    (21, "Upper Primary", "Creative Arts & Sports",       100),
    (22, "Upper Primary", "Home Science",                 100),
    (23, "Upper Primary", "Physical & Health Education",  100),
    # Junior Secondary (Grade 7-9)
    (24, "Junior Sec", "English",                   100),
    (25, "Junior Sec", "Kiswahili",                 100),
    (26, "Junior Sec", "Mathematics",               100),
    (27, "Junior Sec", "Integrated Science",        100),
    (28, "Junior Sec", "Social Studies",            100),
    (29, "Junior Sec", "Pre-Technical Studies",     100),
    (30, "Junior Sec", "Agriculture",               100),
    (31, "Junior Sec", "Creative Arts & Sports",    100),
    (32, "Junior Sec", "Religious Education (CRE)", 100),
    (33, "Junior Sec", "Business Studies",          100),
    (34, "Junior Sec", "Computer Science",          100),
]
seed_data(ws_subj, subj_data)

# 9. DB_Marks  — col layout MUST match SaveAllMarks in modMarks:
# col1=MarkID, col2=StudentID, col3=AdmNo, col4=Term, col5=Year,
# col6=SubjectID, col7=Score, col8=MaxMark, col9=Grade, col10=Remarks, col11=RecordedBy
make_sheet("DB_Marks", [
    "MarkID", "StudentID", "AdmNo", "Term", "Year",
    "SubjectID", "Score", "MaxMark", "Grade", "Remarks", "RecordedBy"
])

# 10. Setup_Grades — CBC grade boundaries per level
ws_gr = make_sheet("Setup_Grades", ["Level", "MinScore", "MaxScore", "Grade", "Remark"])
grades_data = [
    # PP
    ("PP", 0, 39, "Below Expectations", "Needs Improvement"),
    ("PP", 40, 59, "Approaching Expectations", "Fair"),
    ("PP", 60, 79, "Meeting Expectations", "Good"),
    ("PP", 80, 100, "Exceeding Expectations", "Excellent"),
    # Lower Primary
    ("Lower Primary", 0, 39, "Below Expectations", "Needs Improvement"),
    ("Lower Primary", 40, 59, "Approaching Expectations", "Fair"),
    ("Lower Primary", 60, 79, "Meeting Expectations", "Good"),
    ("Lower Primary", 80, 100, "Exceeding Expectations", "Excellent"),
    # Upper Primary
    ("Upper Primary", 0, 39, "Below Expectations", "Needs Improvement"),
    ("Upper Primary", 40, 59, "Approaching Expectations", "Fair"),
    ("Upper Primary", 60, 79, "Meeting Expectations", "Good"),
    ("Upper Primary", 80, 100, "Exceeding Expectations", "Excellent"),
    # Junior Secondary
    ("Junior Sec", 0, 39, "Below Expectations", "Needs Improvement"),
    ("Junior Sec", 40, 59, "Approaching Expectations", "Fair"),
    ("Junior Sec", 60, 79, "Meeting Expectations", "Good"),
    ("Junior Sec", 80, 100, "Exceeding Expectations", "Excellent"),
]
seed_data(ws_gr, grades_data)

# 11. DB_Users — Col order: ID, Username, Password, FullName, Role, Active
ws_usr = make_sheet("DB_Users", ["ID", "Username", "Password", "FullName", "Role", "Active", "Email", "Status"])
users_data = [
    (1, "admin",       "admin123",   "System Administrator", "Admin",       "Yes", "", ""),
    (2, "teacher",     "teacher123", "Default Teacher",      "Teacher",     "Yes", "", ""),
    (3, "headteacher", "head123",    "Head Teacher",         "Headteacher", "Yes", "", ""),
    (4, "bursar",      "bursar123",  "School Bursar",        "Bursar",      "Yes", "", ""),
]
seed_data(ws_usr, users_data)

# 12. DB_SystemLogs
make_sheet("DB_SystemLogs", ["Timestamp", "User", "Action", "Details"])

# 13. DB_Settings — Pre-populated with school info
ws_sett = make_sheet("DB_Settings", ["SettingKey", "SettingValue"])
settings_data = [
    ("SchoolName",    "Magana Primary School"),
    ("SchoolMotto",   "Education for Excellence"),
    ("County",        ""),
    ("SubCounty",     ""),
    ("AcademicYear",  "2026"),
    ("CurrentTerm",   "Term 1"),
    ("Principal",     ""),
    ("ContactPhone",  ""),
    ("ContactEmail",  ""),
    ("POBox",         ""),
]
seed_data(ws_sett, settings_data)

# 14. DB_CBC_Skills
make_sheet("DB_CBC_Skills", [
    "ID", "StudentID", "Term", "Year", "Class",
    "Competency", "Rating"
])

# 15. DB_Physical
make_sheet("DB_Physical", [
    "ID", "StudentID", "Term", "Year",
    "Height_cm", "Weight_kg", "BMI", "Comments"
])

# 16. REPORT OUTPUT (blank, used by report generation)
ws_rpt = wb.create_sheet(title="REPORT OUTPUT")
ws_rpt.sheet_properties.tabColor = "C62828"

# 17. Report_Class
ws_rc = wb.create_sheet(title="Report_Class")
ws_rc.sheet_properties.tabColor = "1565C0"

# 18. Template_ReportCard
ws_tmpl = wb.create_sheet(title="Template_ReportCard")
ws_tmpl.sheet_properties.tabColor = "6A1B9A"

# 19. Dashboard
ws_dash = wb.create_sheet(title="Dashboard")
ws_dash.sheet_properties.tabColor = "2E7D32"

# ── Remove default "Sheet" ──
if "Sheet" in wb.sheetnames:
    del wb["Sheet"]

# ── Move Dashboard to first position ──
wb.move_sheet("Dashboard", offset=-len(wb.sheetnames)+1)

# ── Save ──
wb.save(OUTPUT)
print(f"✅ Workbook created: {OUTPUT}")
print(f"   Sheets: {len(wb.sheetnames)}")
for s in wb.sheetnames:
    print(f"   • {s}")
print(f"\n   📊 Pre-populated data:")
print(f"      • {len(classes_data)} class/stream entries ({len(cbc_classes)} classes × {len(default_streams)} streams)")
print(f"      • {len(subj_data)} CBC curriculum subjects")
print(f"      • {len(fee_data)} fee structure items")
print(f"      • {len(grades_data)} grade boundaries")
print(f"      • {len(users_data)} user accounts")
print(f"      • {len(settings_data)} school settings")
print(f"\n   ⚠️  NO sheet protection applied.")
print(f"   📋 Next: Open in Excel → Alt+F11 → paste AutoSetup → Alt+F8 → Run")
